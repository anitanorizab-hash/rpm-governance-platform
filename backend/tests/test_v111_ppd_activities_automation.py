"""V1.1.1 tests: org detection, dual-layout parsing, activities, composite codes, activity edit, automation."""
import io
import uuid

import openpyxl
from sqlalchemy import select

from app.models.operational.kpi import KPI
from app.models.operational.organisation import Organisation
from app.services import import_service
from app.skills import excel_parsing_skill as parser

PWD = "Password123!"


def _auth(t):
    return {"Authorization": f"Bearer {t}"}


def _admin_token(client, make_admin, email="adm111@moe.gov.my"):
    client.post("/api/v1/auth/register", json={"email": email, "name": "Adm", "password": PWD})
    make_admin(email)
    return client.post("/api/v1/auth/login", json={"email": email, "password": PWD}).json()["access_token"]


def _jpn_book() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PELAN TAKTIKAL JPN(Teras 1&2)"
    ws.append(["RANCANGAN"])
    ws.append(["Bil", "Teras", "KPI", "Sasaran 2026", "Aktiviti Utama", "Milestone (output)",
               "Bahagian - Pegawai", "Aktiviti Sokongan", "Nota Pengiraan"])
    ws.append([1, 1, "JPN KPI statement", "100%", "Main activity", "Milestone X",
               "SPS En. Rao", "Support activity", "nota"])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _ppd_book() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "TERAS 2"
    ws.append(["Plan Taktikal"])
    ws.append(["Bil.", "Pemetaan PPPM 2026-2035", "KPI Daerah", "Sasaran KPI (2026)", "Aktiviti Utama",
               "Milestone", "Peneraju KPI (Sektor/Unit JPN/PPD)", "Penyokong KPI (Sektor/Unit JPN/PPD)",
               "Status Pelaksanaan Aktiviti", "Catatan"])
    ws.append([1, "TS2 S1 I1 KPI3", "District KPI statement", "90%", "PPD main activity", "PPD milestone",
               "SPS", "PPD HEM", "sedang dilaksanakan", "catatan note"])
    # DMU sheet must be skipped structurally
    d = wb.create_sheet("TERAS 2 DMU"); d.append(["random", "stuff"])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# 1. Organisation detection from filename
def test_detect_organisation():
    assert import_service.detect_organisation("PELAN TAKTIKAL JPN.xlsx")["type"] == "JPN"
    ppd = import_service.detect_organisation("07 - PPD KINTA UTARA (TERAS 1-TERAS 7).xlsx")
    assert ppd["type"] == "PPD" and ppd["code"] == "PPD-KINTA-UTARA" and ppd["name"] == "PPD Kinta Utara"


# 2. Dual-layout parsing: JPN columns + PPD columns (Teras from sheet name, DMU skipped)
def test_parse_jpn_and_ppd_layouts():
    jpn = parser.parse_workbook(_jpn_book())["rows"][0]
    assert parser.derive_statement(jpn) == "JPN KPI statement"
    assert jpn["aktiviti_utama"] == "Main activity" and jpn["aktiviti_sokongan"] == "Support activity"
    assert jpn["milestone"] == "Milestone X" and parser.derive_pic_name(jpn) == "SPS En. Rao"

    out = parser.parse_workbook(_ppd_book())
    assert "TERAS 2 DMU" in out["skipped_sheets"]            # structurally skipped
    ppd = out["rows"][0]
    assert ppd["teras"] == 2                                  # from sheet title
    assert parser.derive_code(ppd) == "TS2 S1 I1 KPI3"        # from Pemetaan PPPM
    assert ppd["status_pelaksanaan"] == "sedang dilaksanakan" and ppd["catatan"] == "catatan note"


# 3. Import creates Aktiviti Utama + Sokongan + Milestone, and the PPD organisation
def test_import_creates_activities_and_ppd_org(client, make_admin, db_session):
    tok = _admin_token(client, make_admin)
    r = client.post("/api/v1/imports/execute", headers=_auth(tok),
                    files={"file": ("05 - PPD KERIAN (TERAS 1-TERAS 7).xlsx", _ppd_book())})
    assert r.status_code == 200 and r.json()["organisation_name"] == "PPD Kerian"
    org = db_session.scalar(select(Organisation).where(Organisation.code == "PPD-KERIAN"))
    assert org is not None and org.type == "PPD"
    kpi = db_session.scalar(select(KPI).where(KPI.organisation_id == org.id))
    detail = client.get(f"/api/v1/kpis/{kpi.id}", headers=_auth(tok)).json()
    types = {a["type"] for a in detail["activities"]}
    assert "utama" in types and "sokongan" in types
    assert any(a["milestone"] for a in detail["activities"])


# 4. Composite unique: same KPI code can exist under two organisations
def test_same_code_two_organisations(db_session):
    from app.repositories.import_repository import ImportRepository
    repo = ImportRepository(db_session)
    jpn = repo.get_or_create_organisation(code="JPN", name="JPN", type="JPN")
    ppd = repo.get_or_create_organisation(code="PPD-X", name="PPD X", type="PPD", parent_id=jpn.id)
    repo.create_kpi(code="TS1 S1 I1 KPI1", organisation_id=jpn.id, statement="a")
    repo.create_kpi(code="TS1 S1 I1 KPI1", organisation_id=ppd.id, statement="b")  # must NOT clash
    db_session.commit()
    n = db_session.scalars(select(KPI).where(KPI.code == "TS1 S1 I1 KPI1")).all()
    assert len(n) == 2


# 5. Activity progress edit (operational, audited, not amendment-gated)
def test_activity_progress_update(client, make_admin, db_session):
    tok = _admin_token(client, make_admin)
    client.post("/api/v1/imports/execute", headers=_auth(tok),
                files={"file": ("PELAN TAKTIKAL JPN.xlsx", _jpn_book())})
    kpi = db_session.scalar(select(KPI).where(KPI.statement == "JPN KPI statement"))
    detail = client.get(f"/api/v1/kpis/{kpi.id}", headers=_auth(tok)).json()
    aid = detail["activities"][0]["id"]
    r = client.patch(f"/api/v1/kpis/{kpi.id}/activities/{aid}", headers=_auth(tok),
                     json={"status": "In progress", "remarks": "on track"})
    assert r.status_code == 200
    updated = next(a for a in r.json()["activities"] if a["id"] == aid)
    assert updated["status"] == "In progress" and updated["remarks"] == "on track"
    audit = client.get("/api/v1/audit/logs", headers=_auth(tok)).json()
    assert any(a["action"] == "activity_update" for a in audit)


# 6. Automation generates DRAFTS only (HITL-safe), skips recipients without a valid email
def test_automation_generates_drafts(client, make_admin, db_session):
    tok = _admin_token(client, make_admin)
    client.post("/api/v1/imports/execute", headers=_auth(tok),
                files={"file": ("PELAN TAKTIKAL JPN.xlsx", _jpn_book())})
    r = client.post("/api/v1/automation/run", headers=_auth(tok),
                    json={"types": ["monthly_report", "missing_info"], "limit": 5})
    assert r.status_code == 200
    body = r.json()
    assert body["generated"].get("monthly_report") == 1
    assert "skipped_no_valid_email" in body          # imported PICs have no email → skipped
    # a draft report now exists (pending HITL approval), not auto-issued
    reports = client.get("/api/v1/reports", headers=_auth(tok)).json()
    assert any(rep["status"] == "draft" for rep in reports)


# 7. Automation is admin-only
def test_automation_requires_admin(client):
    client.post("/api/v1/auth/register", json={"email": "plain111@moe.gov.my", "name": "P", "password": PWD})
    tok = client.post("/api/v1/auth/login", json={"email": "plain111@moe.gov.my", "password": PWD}).json()["access_token"]
    r = client.post("/api/v1/automation/run", headers=_auth(tok), json={"types": ["monthly_report"]})
    assert r.status_code == 403
