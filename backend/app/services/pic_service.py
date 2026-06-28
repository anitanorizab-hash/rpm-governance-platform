"""PIC service (V1.1.1 PIC Directory) — admin-managed PICs, KPI assignment, Excel bulk import/export.

Admin-only. Audited. Soft delete. Setting a PIC email here makes notification/report automation use
it automatically (automation reads kpi.pic.email). Does not touch the HITL approval workflow.
"""
from __future__ import annotations

import io

import openpyxl
from sqlalchemy.orm import Session

from app.core.audit import AuditContext
from app.repositories.organisation_repository import OrganisationRepository
from app.repositories.pic_repository import PICRepository
from app.services.audit_service import AuditService

ADMIN_ROLES = {"super_admin", "jpn_admin"}


class PICPermissionError(Exception):
    pass


class PICService:
    def __init__(self, db: Session):
        self.db = db
        self.repo = PICRepository(db)
        self.orgs = OrganisationRepository(db)
        self.audit = AuditService(db)

    def _auth(self, current_user):
        if not (set(current_user.role_names) & ADMIN_ROLES):
            raise PICPermissionError("PIC management is limited to Super Admin / JPN Admin.")

    def list(self, current_user, **filters):
        self._auth(current_user)
        return self.repo.list(**filters)

    def get(self, current_user, pic_id):
        self._auth(current_user)
        return self.repo.get(pic_id)

    def create(self, *, current_user, data: dict, context: AuditContext | None = None):
        self._auth(current_user)
        p = self.repo.create(name=data["name"], email=data.get("email"), sector=data.get("department"),
                             organisation_id=data.get("organisation_id"), active=data.get("active", True))
        self.audit.record(entity_type="pic", entity_id=p.id, action="pic_create", actor_id=current_user.id,
                          after={"name": p.name, "email": p.email}, context=context, commit=False)
        self.db.commit()
        return self.repo.get(p.id)

    def update(self, *, current_user, pic_id, fields: dict, context: AuditContext | None = None):
        self._auth(current_user)
        p = self.repo.get(pic_id)
        if not p:
            return None
        changes = {}
        for key, attr in (("name", "name"), ("email", "email"), ("department", "sector"),
                          ("organisation_id", "organisation_id"), ("active", "active")):
            if fields.get(key) is not None:
                setattr(p, attr, fields[key]); changes[key] = fields[key]
        self.db.flush()
        self.audit.record(entity_type="pic", entity_id=p.id, action="pic_update", actor_id=current_user.id,
                          after=changes, context=context, commit=False)
        self.db.commit()
        return self.repo.get(p.id)

    def delete(self, *, current_user, pic_id, context: AuditContext | None = None):
        self._auth(current_user)
        p = self.repo.get(pic_id)
        if not p:
            return None
        self.repo.soft_delete(p)
        self.audit.record(entity_type="pic", entity_id=p.id, action="pic_soft_delete",
                          actor_id=current_user.id, context=context, commit=False)
        self.db.commit()
        return True

    def assign_kpis(self, *, current_user, pic_id, kpi_ids, context: AuditContext | None = None):
        self._auth(current_user)
        p = self.repo.get(pic_id)
        if not p:
            return None
        n = self.repo.assign_kpis(pic_id, kpi_ids)
        self.audit.record(entity_type="pic", entity_id=pic_id, action="pic_assign_kpis",
                          actor_id=current_user.id, after={"assigned": n}, context=context, commit=False)
        self.db.commit()
        return {"pic_id": pic_id, "assigned": n}

    # ----- bulk export / import (Excel) -----
    def export_excel(self, current_user) -> bytes:
        self._auth(current_user)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PIC Directory"
        ws.append(["Name", "Email", "Organisation", "Department", "Status", "Assigned KPIs"])
        for p in self.repo.list(include_deleted=False):
            codes = ", ".join(k.code for k in self.repo.assigned_kpis(p.id))
            ws.append([p.name, p.email or "", (p.organisation.code if p.organisation else ""),
                       p.sector or "", "Active" if p.active else "Inactive", codes])
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def import_excel(self, *, current_user, file_bytes: bytes, context: AuditContext | None = None) -> dict:
        self._auth(current_user)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            return {"created": 0, "updated": 0}
        header = [str(c or "").strip().lower() for c in rows[0]]

        def col(name):
            return header.index(name) if name in header else -1
        i_name, i_email, i_org, i_dept, i_status = (col("name"), col("email"), col("organisation"),
                                                    col("department"), col("status"))
        created = updated = 0
        for r in rows[1:]:
            def cell(i):
                return (str(r[i]).strip() if 0 <= i < len(r) and r[i] not in (None, "") else None)
            name = cell(i_name)
            if not name:
                continue
            email, dept = cell(i_email), cell(i_dept)
            org = self.orgs.get_by_code(cell(i_org)) if cell(i_org) else None
            active = (cell(i_status) or "active").lower() != "inactive"
            existing = self.repo.get_by_email(email) if email else None
            if existing:
                existing.name = name
                if dept:
                    existing.sector = dept
                if org:
                    existing.organisation_id = org.id
                existing.active = active
                updated += 1
            else:
                self.repo.create(name=name, email=email, sector=dept,
                                 organisation_id=(org.id if org else None), active=active)
                created += 1
        self.audit.record(entity_type="pic", action="pic_import", actor_id=current_user.id,
                          after={"created": created, "updated": updated}, context=context, commit=False)
        self.db.commit()
        return {"created": created, "updated": updated}
