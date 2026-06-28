"""Excel parsing skill (CP6; V1.1.1 dual-layout) — reusable, deterministic, independently testable.

Reads a Pelan Taktikal workbook (JPN or PPD layout) into canonical row dicts via a synonym-based,
longest-match-first column map. Auto-detects the header row; derives Teras from the sheet name when
there is no Teras column (PPD); structurally selects KPI Tactical Plan sheets and skips
reference/DMU/SPP/summary/cover sheets. Used ONLY by the one-time import pipeline.
"""
from __future__ import annotations

import io
import re

import openpyxl

# Canonical field → accepted header prefixes (normalised, lower-case). A header matches a synonym
# when it STARTS WITH the synonym; the LONGEST matching synonym wins, so "kpi nasional" beats "kpi".
COLUMN_MAP: dict[str, list[str]] = {
    "teras": ["teras"],
    "pemetaan": ["pemetaan pppm", "pemetaan"],          # PPD: carries the national mapping code
    "kpi_code": ["kpi code", "kod kpi", "kod"],
    "kpi_daerah": ["kpi daerah"],
    "kpi_negeri": ["kpi negeri"],
    "kpi_nasional": ["kpi nasional"],
    "kpi_statement": ["kpi statement", "penyataan kpi", "kpi"],          # JPN single "KPI" column
    "indicator": ["indikator", "indicator", "petunjuk"],
    "target": ["sasaran kpi", "sasaran 2026", "sasaran", "target"],
    "tov": ["tov kpi", "tov", "pencapaian 2025", "baseline"],
    "pic_lead": ["bahagian - pegawai", "bahagian-pegawai", "peneraju kpi", "pegawai",
                 "pic name", "nama pic"],
    "pic_support": ["penyokong kpi"],
    "pic_email": ["emel pic", "pic email", "email", "emel"],
    "aktiviti_utama": ["aktiviti utama"],
    "aktiviti_sokongan": ["aktiviti sokongan"],
    "milestone": ["milestone"],
    "nota_pengiraan": ["nota pengiraan"],
    "catatan": ["catatan"],
    "status_pelaksanaan": ["status pelaksanaan"],
    "department": ["bahagian", "sektor", "sector", "department"],
    "amount": ["jumlah (rm)", "jumlah", "amaun", "amount"],
}

# Flattened (canonical, synonym) pairs, longest synonym first → specific headers win.
_SYNONYMS = sorted(
    ((canonical, syn) for canonical, syns in COLUMN_MAP.items() for syn in syns),
    key=lambda pair: len(pair[1]), reverse=True,
)

# A sheet is a KPI Tactical Plan sheet only if its header carries a KPI-bearing column…
_KPI_FIELDS = {"kpi_statement", "kpi_daerah", "kpi_negeri", "kpi_nasional", "kpi_code"}
# …and at least one operational column.
_OPERATIONAL_FIELDS = {"target", "aktiviti_utama", "milestone", "tov"}
# Sheets that never hold KPI rows (fast-path skip; structural check still applies otherwise).
_SKIP_SHEET = re.compile(r"reference|senarai|dmu|spp|cover|kandungan|arahan|instruction|panduan|"
                         r"summary|ringkasan|rumusan", re.I)

MANDATORY_FIELDS = ("kpi_statement", "pic_name", "teras")   # email is captured manually (not in source)


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def teras_from_title(title: str) -> int | None:
    """Derive Teras number from a sheet title like 'TERAS 1' / 'Teras 3&4' → 1 / 3."""
    m = re.search(r"teras\s*([1-7])", title or "", re.I)
    return int(m.group(1)) if m else None


def _build_header_index(header_cells: list) -> dict[int, str]:
    """Map column index → canonical field for a candidate header row (longest synonym wins)."""
    mapping: dict[int, str] = {}
    for idx, cell in enumerate(header_cells):
        text = _norm(cell)
        if not text:
            continue
        for canonical, syn in _SYNONYMS:
            if canonical in mapping.values():
                continue
            if text == syn or text.startswith(syn):
                mapping[idx] = canonical
                break
    return mapping


def _find_header_row(rows: list[list], scan: int = 25) -> tuple[int, dict[int, str]]:
    best_row, best_map = -1, {}
    for i, row in enumerate(rows[:scan]):
        m = _build_header_index(row)
        if len(m) > len(best_map):
            best_row, best_map = i, m
    return best_row, best_map


def _is_kpi_sheet(header_map: dict[int, str]) -> bool:
    fields = set(header_map.values())
    return bool(fields & _KPI_FIELDS) and bool(fields & _OPERATIONAL_FIELDS)


def parse_sheet(rows: list[list], teras_hint: int | None = None) -> list[dict]:
    """Parse one sheet's rows into canonical row dicts. Returns [] for non-KPI sheets."""
    header_row, header_map = _find_header_row(rows)
    if header_row < 0 or not _is_kpi_sheet(header_map):
        return []
    parsed: list[dict] = []
    for row in rows[header_row + 1:]:
        record = {field: None for field in COLUMN_MAP}
        has_value = False
        for idx, field in header_map.items():
            val = row[idx] if idx < len(row) else None
            if val not in (None, ""):
                record[field] = val
                has_value = True
        if teras_hint is not None and not record.get("teras"):
            record["teras"] = teras_hint
        if has_value:
            parsed.append(record)
    return parsed


def parse_workbook(source, sheet_names: list[str] | None = None) -> dict:
    """Parse a workbook (path or bytes) → {rows, sheets, skipped_sheets}.

    Selects KPI Tactical Plan sheets structurally; derives Teras from the sheet title when absent.
    """
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True, read_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True, read_only=True)

    all_rows: list[dict] = []
    used, skipped = 0, []
    for ws in wb.worksheets:
        if sheet_names is not None and ws.title not in sheet_names:
            continue
        if sheet_names is None and _SKIP_SHEET.search(ws.title):
            skipped.append(ws.title)
            continue
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        sheet_rows = parse_sheet(rows, teras_hint=teras_from_title(ws.title))
        if sheet_rows:
            used += 1
            all_rows.extend(sheet_rows)
        else:
            skipped.append(ws.title)
    return {"rows": all_rows, "sheets": used, "skipped_sheets": skipped}


# ---- canonical row helpers (used by the import service) ----
_CODE_RE = re.compile(r"TS\s*\d+\s*S\d+\s*I\d+\s*KPI\s*\d+", re.I)


def derive_statement(record: dict) -> str | None:
    """Best KPI statement across layouts: JPN 'KPI' or PPD Daerah → Negeri → Nasional."""
    for f in ("kpi_statement", "kpi_daerah", "kpi_negeri", "kpi_nasional"):
        v = record.get(f)
        if v not in (None, ""):
            return str(v).strip()
    return None


def derive_code(record: dict) -> str | None:
    """Explicit KPI code, else extract a 'TSx Sy Iz KPIn' pattern from the national KPI cell."""
    if record.get("kpi_code"):
        return str(record["kpi_code"]).strip()
    for f in ("pemetaan", "kpi_nasional", "kpi_negeri", "kpi_statement"):
        v = record.get(f)
        if v:
            m = _CODE_RE.search(str(v))
            if m:
                return re.sub(r"\s+", " ", m.group(0)).strip()
    return None


def derive_pic_name(record: dict) -> str | None:
    v = record.get("pic_lead")
    return str(v).strip() if v not in (None, "") else None


def missing_fields(record: dict) -> list[str]:
    """Return missing mandatory fields (statement, pic_name, teras). Email is manual."""
    miss = []
    if not derive_statement(record):
        miss.append("kpi_statement")
    if not derive_pic_name(record):
        miss.append("pic_name")
    if not record.get("teras"):
        miss.append("teras")
    return miss
